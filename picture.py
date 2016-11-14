# encoding: utf-8
from os import listdir
from openpyxl import load_workbook
from itertools import permutations
from random import uniform


class DataIter:

    def __init__(self, common_name, scientific_name, is_recite):
        self._common_name = common_name
        self._scientific_name = scientific_name
        self._is_recite = is_recite

    @property
    def common_name(self):
        return self._common_name

    @property
    def scientific_name(self):
        return self._scientific_name

    @property
    def is_recite(self):
        return self._common_name

    @is_recite.setter
    def is_recite(self, value):
        pass

    @scientific_name.setter
    def scientific_name(self, value):
        pass

    @common_name.setter
    def common_name(self, value):
        pass


class DataLoad:

    picture_type = ['.jpg', '.png', '.bmp']

    # 此处如果要广泛应用数值需要修改
    def __init__(self, excel_name='000.xlsx', p_name='000', picture_num=200): 
        self.excel = load_workbook(excel_name)
        self.p_name = p_name
        self.picture_num = picture_num
        self._list = listdir('.')

    def get_picture_name(self, no):
        for suffix in self.picture_type:
            if no + suffix in self._list:
                return no+suffix

    def get_cell_value(self, name='000', x=0, y=0):
        ws = self.excel.get_sheet_by_name(name)
        n = ws.cell(row=x, column=y)
        return n.value

    def view_list(self, num=200):
        for row in range(1, num+1):
            name = self.get_picture_name(self.get_cell_value(x=row, y=1))
            if name:
                yield DataIter(self.get_cell_value(x=row, y=2), self.get_cell_value(x=row, y=3),\
                               self.get_cell_value(x=row, y=4)), name

    def test_list(self, num=200):
        perm = permutations([i for i in range(1, num+1)])
        for i in range(int(uniform(num*num*num, num*num*num*num))):
            next(perm)
        for row in reversed(next(perm)):
            name = self.get_picture_name(self.get_cell_value(x=row, y=1))
            if name:
                yield DataIter(self.get_cell_value(x=row, y=2), self.get_cell_value(x=row, y=3),\
                               self.get_cell_value(x=row, y=4)), name


def main_example():
    data_load = DataLoad()
    test_iter = data_load.test_list()

    '''
    第一种使用方式
    data_iter:保存各种名字的对象
    picture_name:保存对应图片名字的变量
    '''
    for data_iter, picture_name in test_iter:
        print(data_iter,picture_name)
    '''
    第二种使用方式
    每次调用next会获得下一个植物的数据
    '''
    data_iter, picture_name = next(test_iter)
    print(data_iter, picture_name)


<<<<<<< HEAD

=======
>>>>>>> origin/master
